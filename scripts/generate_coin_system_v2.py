# -*- coding: utf-8 -*-
"""海外短剧金币系统分析 - 极简单页版"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ===== 样式 =====
TITLE_FONT = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
TITLE_FILL = PatternFill('solid', fgColor='2F5496')

# 四大色块
BLOCK_FILLS = {
    '获取': PatternFill('solid', fgColor='4472C4'),   # 蓝
    '消耗': PatternFill('solid', fgColor='C00000'),   # 红
    '任务': PatternFill('solid', fgColor='548235'),   # 绿
    '变现': PatternFill('solid', fgColor='BF8F00'),   # 黄
}
BLOCK_FONTS = {k: Font(name='微软雅黑', size=11, bold=True, color='FFFFFF') for k in BLOCK_FILLS}

COL_FONT = Font(name='微软雅黑', size=10, bold=True, color='1F3864')
DATA_FONT = Font(name='微软雅黑', size=10)
THIN = Border(
    left=Side('thin', 'B4C7E7'), right=Side('thin', 'B4C7E7'),
    top=Side('thin', 'B4C7E7'), bottom=Side('thin', 'B4C7E7'),
)
C = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 浅色背景（交替）
LIGHT = {
    '获取': PatternFill('solid', fgColor='D6E4F0'),
    '消耗': PatternFill('solid', fgColor='FCE4E4'),
    '任务': PatternFill('solid', fgColor='E2EFDA'),
    '变现': PatternFill('solid', fgColor='FFF2CC'),
}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '金币系统对比'
ws.sheet_view.showGridLines = False

# ===== 列定义 =====
# (列名, 所属色块, 宽度)
columns = [
    ('产品名称',        None,   18),
    ('变现模式',        None,   10),
    # 获取 - 蓝
    ('签到\n(金币/天)',  '获取', 12),
    ('激励视频\n(金币/次)', '获取', 12),
    ('激励视频\n(次/天)', '获取', 10),
    ('分享\n(金币/次)',  '获取', 10),
    ('邀请\n(金币/人)',  '获取', 10),
    ('每日产出\n上限',   '获取', 12),
    # 消耗 - 红
    ('解锁1集\n(金币)',  '消耗', 12),
    ('阶梯价',          '消耗', 10),
    ('其他消耗\n场景',   '消耗', 18),
    # 任务 - 绿
    ('日常任务\n(类型&奖励)', '任务', 22),
    ('社交任务\n(类型&奖励)', '任务', 22),
    ('成长/成就\n(类型&奖励)', '任务', 22),
    # 变现 - 黄
    ('广告点位\n(类型)', '变现', 18),
    ('1$=\n金币数',     '变现', 10),
    ('提现\n(有/无)',   '变现', 10),
    ('提现门槛\n($)',   '变现', 10),
    ('核心亮点/备注',    None,   30),
]

# ===== 标题行 =====
total_cols = len(columns)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
ws['A1'] = '海外短剧金币系统对比表'
ws['A1'].font = TITLE_FONT
ws['A1'].fill = TITLE_FILL
ws['A1'].alignment = C
ws.row_dimensions[1].height = 32

# ===== 色块行 (row 2) =====
col = 1
# 产品名称 + 变现模式 合并
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
ws.cell(row=2, column=1, value='基本信息').font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
ws.cell(row=2, column=1).fill = PatternFill('solid', fgColor='595959')
ws.cell(row=2, column=1).alignment = C

# 按色块合并
block_ranges = {}
start = 3
for i, (name, block, w) in enumerate(columns):
    col_idx = i + 1
    if block and (i == 0 or columns[i-1][1] != block):
        start = col_idx
    if block and (i == len(columns)-1 or columns[i+1][1] != block):
        ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=col_idx)
        cell = ws.cell(row=2, column=start, value=block)
        cell.font = BLOCK_FONTS[block]
        cell.fill = BLOCK_FILLS[block]
        cell.alignment = C

# 核心亮点/备注
last_col = total_cols
ws.cell(row=2, column=last_col, value='备注')
ws.cell(row=2, column=last_col).font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
ws.cell(row=2, column=last_col).fill = PatternFill('solid', fgColor='595959')
ws.cell(row=2, column=last_col).alignment = C

ws.row_dimensions[2].height = 22

# ===== 列名行 (row 3) =====
for i, (name, block, w) in enumerate(columns):
    col_idx = i + 1
    cell = ws.cell(row=3, column=col_idx, value=name)
    cell.font = COL_FONT
    cell.fill = LIGHT.get(block, PatternFill('solid', fgColor='F2F2F2'))
    cell.alignment = C
    cell.border = THIN
    ws.column_dimensions[get_column_letter(col_idx)].width = w

ws.row_dimensions[3].height = 38

# ===== 预填产品 + 空行 =====
products = [
    'ReelShort', 'DramaBox', 'ShortTV', 'FlexTV', 'GoodShort',
    'MiniShorts', 'DramaWave', 'TopShorts', 'DreameShort', 'Episode',
    'ShortMax', 'ViuShort', 'MangaToon', 'Pocket FM', 'DramaSala',
    '', '', '', '', '', '', '', '', '', '',
    '', '', '', '', '', '', '', '', '', '',
    '', '', '', '', '', '', '', '', '', '',
    '', '', '', '', '', '', '', '', '', '',
]

for row_idx, product in enumerate(products, start=4):
    ws.cell(row=row_idx, column=1, value=product if product else None)
    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = THIN
        cell.font = DATA_FONT
        cell.alignment = C
        # 交替行颜色
        block = columns[col_idx - 1][1]
        if block:
            if row_idx % 2 == 0:
                cell.fill = LIGHT[block]
ws.freeze_panes = 'C4'

# ===== 下拉菜单 =====
dv_mode = DataValidation(type='list', formula1='"IAA,IAAP,IAP"', allow_blank=True)
dv_mode.add(f'B4:B53')
ws.add_data_validation(dv_mode)

dv_yn = DataValidation(type='list', formula1='"是,否"', allow_blank=True)
dv_yn.add(f'J4:J53')  # 阶梯价
ws.add_data_validation(dv_yn)

dv_tw = DataValidation(type='list', formula1='"有,无"', allow_blank=True)
dv_tw.add(f'Q4:Q53')  # 提现
ws.add_data_validation(dv_tw)

# ===== 第二个Sheet: 任务详情（补充用，非必须） =====
ws2 = wb.create_sheet('任务详情（补充）')
ws2.sheet_view.showGridLines = False
ws2.merge_cells('A1:H1')
ws2['A1'] = '任务详情补充表 — 每个产品的每个任务单独一行'
ws2['A1'].font = TITLE_FONT
ws2['A1'].fill = TITLE_FILL
ws2['A1'].alignment = C
ws2.row_dimensions[1].height = 30

task_headers = ['产品名称', '任务类型', '任务名称', '奖励(金币)', '完成条件', '可重复', '刷新周期', '备注']
for col, h in enumerate(task_headers, 1):
    cell = ws2.cell(row=2, column=col, value=h)
    cell.font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='548235')
    cell.alignment = C
    cell.border = THIN

ws2.row_dimensions[2].height = 30
task_widths = [16, 14, 22, 12, 25, 10, 12, 20]
for i, w in enumerate(task_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

dv_task = DataValidation(type='list', formula1='"日常任务,激励广告,社交任务,成长任务,成就任务,活动任务,新手任务"', allow_blank=True)
dv_task.add('B3:B200')
ws2.add_data_validation(dv_task)

dv_rep = DataValidation(type='list', formula1='"可重复,不可重复,每日1次,每周1次"', allow_blank=True)
dv_rep.add('F3:F200')
ws2.add_data_validation(dv_rep)

for row in range(3, 203):
    for col in range(1, 9):
        cell = ws2.cell(row=row, column=col)
        cell.border = THIN
        cell.font = DATA_FONT
        cell.alignment = C

ws2.freeze_panes = 'A3'

# ===== 保存 =====
output_path = '海外短剧金币系统对比表.xlsx'
wb.save(output_path)
print(f'Done: {output_path}')
