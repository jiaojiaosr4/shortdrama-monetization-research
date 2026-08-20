# -*- coding: utf-8 -*-
"""海外短剧金币系统分析模板生成器"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

# ===== 样式定义 =====
TITLE_FONT = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
TITLE_FILL = PatternFill('solid', fgColor='2F5496')
HEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='4472C4')
SUB_HEADER_FONT = Font(name='微软雅黑', size=10, bold=True, color='1F3864')
SUB_HEADER_FILL = PatternFill('solid', fgColor='D6E4F0')
NOTE_FONT = Font(name='微软雅黑', size=9, italic=True, color='808080')
DATA_FONT = Font(name='微软雅黑', size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='B4C7E7'),
    right=Side(style='thin', color='B4C7E7'),
    top=Side(style='thin', color='B4C7E7'),
    bottom=Side(style='thin', color='B4C7E7'),
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
HIGHLIGHT_FILL = PatternFill('solid', fgColor='FFF2CC')
SECTION_FILL = PatternFill('solid', fgColor='E2EFDA')

wb = openpyxl.Workbook()

# ===== Sheet 0: 使用说明 =====
ws0 = wb.active
ws0.title = '使用说明'
ws0.sheet_view.showGridLines = False
ws0.merge_cells('A1:F1')
ws0['A1'] = '海外短剧金币系统分析模板'
ws0['A1'].font = TITLE_FONT
ws0['A1'].fill = TITLE_FILL
ws0['A1'].alignment = CENTER
ws0.row_dimensions[1].height = 40

instructions = [
    ('Sheet名称', '用途', '使用方法'),
    ('产品总览矩阵', '50个IAAP产品的横向对比总表', '每个产品一行，填入基础信息和金币系统概况，用于快速筛选和对比'),
    ('金币获取机制', '采集每个产品的金币产出入口和奖励', '按产品逐行填写获取方式、单次奖励、频次上限等'),
    ('金币消耗机制', '采集每个产品的金币消耗场景和兑换比例', '重点记录解锁一集所需金币、阶梯价设计等'),
    ('任务系统拆解', '按任务类型详细拆解（重点Sheet）', '每个产品x每个任务类型一行，记录奖励/门槛/刷新等'),
    ('经济模型计算', '自动计算关键经济指标', '填入产出和消耗数据后查看自动计算结果'),
    ('变现路径设计', '广告变现+内购+提现的设计对比', '记录广告点位、内购档位、提现门槛等'),
    ('典型样本深度拆解', '3-5个重点产品的深度分析', '选头部/腰部/差异化产品做深度拆解'),
    ('', '', ''),
    ('调研优先级', '建议顺序', ''),
    ('第一步', '先填「产品总览矩阵」筛选出有金币系统的产品', ''),
    ('第二步', '从有金币系统的产品中选3-5个填「典型样本深度拆解」', ''),
    ('第三步', '对典型样本逐一填写「获取/消耗/任务/变现」四个Sheet', ''),
    ('第四步', '汇总数据到「经济模型计算」验证平衡性', ''),
]

for row_idx, row_data in enumerate(instructions, start=3):
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws0.cell(row=row_idx, column=col_idx, value=val)
        cell.border = THIN_BORDER
        cell.alignment = LEFT if col_idx <= 2 else LEFT
        if row_idx == 3:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
        else:
            cell.font = DATA_FONT

ws0.column_dimensions['A'].width = 20
ws0.column_dimensions['B'].width = 40
ws0.column_dimensions['C'].width = 60

# ===== Sheet 1: 产品总览矩阵 =====
ws1 = wb.create_sheet('产品总览矩阵')
ws1.sheet_view.showGridLines = False
ws1.merge_cells('A1:O1')
ws1['A1'] = 'IAAP短剧产品 — 金币系统总览矩阵（50款）'
ws1['A1'].font = TITLE_FONT
ws1['A1'].fill = TITLE_FILL
ws1['A1'].alignment = CENTER
ws1.row_dimensions[1].height = 35

overview_headers = [
    '序号', '产品名称', '厂商/地区', '变现模式\n(IAA/IAAP/IAP)',
    '是否有\n金币系统', '是否有\n提现机制', '是否有\n会员体系',
    '金币获取\n入口数量', '金币消耗\n场景数量', '任务类型\n数量',
    '广告点位\n数量', '每日产出\n上限(金币)', '解锁1集\n所需金币',
    '核心特征\n一句话', '备注'
]

for col, h in enumerate(overview_headers, 1):
    cell = ws1.cell(row=2, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

ws1.row_dimensions[2].height = 40

# 添加50行空行
for row in range(3, 53):
    ws1.cell(row=row, column=1, value=row - 2)
    for col in range(1, 16):
        cell = ws1.cell(row=row, column=col)
        cell.border = THIN_BORDER
        cell.font = DATA_FONT
        cell.alignment = CENTER

# 列宽
col_widths_1 = [6, 20, 15, 14, 10, 10, 10, 12, 12, 10, 10, 14, 14, 30, 20]
for i, w in enumerate(col_widths_1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# 数据验证 - 下拉菜单
dv_mode = DataValidation(type='list', formula1='"IAA,IAAP,IAP"', allow_blank=True)
dv_mode.add('D3:D52')
ws1.add_data_validation(dv_mode)

dv_yesno = DataValidation(type='list', formula1='"有,无,不确定"', allow_blank=True)
dv_yesno.add('E3:E52')
dv_yesno.add('F3:F52')
dv_yesno.add('G3:G52')
ws1.add_data_validation(dv_yesno)

ws1.freeze_panes = 'C3'

# ===== Sheet 2: 金币获取机制 =====
ws2 = wb.create_sheet('金币获取机制')
ws2.sheet_view.showGridLines = False
ws2.merge_cells('A1:N1')
ws2['A1'] = '金币获取机制 — 产出端数据采集'
ws2['A1'].font = TITLE_FONT
ws2['A1'].fill = TITLE_FILL
ws2['A1'].alignment = CENTER
ws2.row_dimensions[1].height = 35

acquire_headers = [
    '序号', '产品名称',
    '激励视频\n奖励(金币)', '激励视频\n每日次数',
    '签到奖励\n(金币)', '签到连续\n递增机制',
    '分享奖励\n(金币)', '分享每日\n次数上限',
    '邀请奖励\n(金币/人)', '邀请上限\n(人/日)',
    '追剧任务\n奖励(金币)', '成就任务\n奖励(金币)',
    '每日总产出\n上限(金币)', '会员倍率\n(1.5x/2x等)'
]

for col, h in enumerate(acquire_headers, 1):
    cell = ws2.cell(row=2, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

ws2.row_dimensions[2].height = 40

for row in range(3, 53):
    ws2.cell(row=row, column=1, value=row - 2)
    for col in range(1, 15):
        cell = ws2.cell(row=row, column=col)
        cell.border = THIN_BORDER
        cell.font = DATA_FONT
        cell.alignment = CENTER

col_widths_2 = [6, 20, 14, 12, 14, 14, 14, 12, 14, 12, 14, 14, 14, 14]
for i, w in enumerate(col_widths_2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = 'C3'

# ===== Sheet 3: 金币消耗机制 =====
ws3 = wb.create_sheet('金币消耗机制')
ws3.sheet_view.showGridLines = False
ws3.merge_cells('A1:M1')
ws3['A1'] = '金币消耗机制 — 消耗端数据采集'
ws3['A1'].font = TITLE_FONT
ws3['A1'].fill = TITLE_FILL
ws3['A1'].alignment = CENTER
ws3.row_dimensions[1].height = 35

consume_headers = [
    '序号', '产品名称',
    '消耗场景\n数量', '场景1\n(解锁剧集)',
    '解锁1集\n所需金币', '是否有\n阶梯定价',
    '阶梯定价\n说明', '场景2\n(抽奖/转盘)',
    '抽奖单次\n所需金币', '场景3\n(虚拟道具)',
    '道具价格\n区间(金币)', '是否有\n付费旁路',
    '付费解锁\n1集价格($)'
]

for col, h in enumerate(consume_headers, 1):
    cell = ws3.cell(row=2, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

ws3.row_dimensions[2].height = 40

for row in range(3, 53):
    ws3.cell(row=row, column=1, value=row - 2)
    for col in range(1, 14):
        cell = ws3.cell(row=row, column=col)
        cell.border = THIN_BORDER
        cell.font = DATA_FONT
        cell.alignment = CENTER

col_widths_3 = [6, 20, 10, 16, 14, 10, 25, 16, 14, 14, 16, 10, 14]
for i, w in enumerate(col_widths_3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

dv_yesno3 = DataValidation(type='list', formula1='"是,否,不确定"', allow_blank=True)
dv_yesno3.add('F3:F52')
dv_yesno3.add('L3:L52')
ws3.add_data_validation(dv_yesno3)

ws3.freeze_panes = 'C3'

# ===== Sheet 4: 任务系统拆解（重点） =====
ws4 = wb.create_sheet('任务系统拆解')
ws4.sheet_view.showGridLines = False
ws4.merge_cells('A1:I1')
ws4['A1'] = '任务系统拆解 — 按任务类型逐条采集（重点Sheet）'
ws4['A1'].font = TITLE_FONT
ws4['A1'].fill = TITLE_FILL
ws4['A1'].alignment = CENTER
ws4.row_dimensions[1].height = 35

# 任务类型说明行
ws4.merge_cells('A2:I2')
ws4['A2'] = '任务类型参考：①日常任务 ②激励广告任务 ③社交任务 ④成长任务 ⑤成就任务 ⑥活动任务 ⑦新手任务  ｜  每个产品的每个任务类型单独一行'
ws4['A2'].font = NOTE_FONT
ws4['A2'].alignment = LEFT
ws4.row_dimensions[2].height = 25

task_headers = [
    '序号', '产品名称', '任务类型', '任务名称/描述',
    '奖励金币', '完成门槛/条件', '可重复性', '刷新周期', '备注'
]

for col, h in enumerate(task_headers, 1):
    cell = ws4.cell(row=3, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

ws4.row_dimensions[3].height = 35

# 预填模板行 - 每个产品7种任务类型
task_types = ['日常任务', '激励广告任务', '社交任务', '成长任务', '成就任务', '活动任务', '新手任务']
row_num = 4
for product_idx in range(1, 6):  # 先给5个产品预填模板
    for task_type in task_types:
        ws4.cell(row=row_num, column=1, value=row_num - 3)
        ws4.cell(row=row_num, column=2, value=f'产品{product_idx}')
        ws4.cell(row=row_num, column=3, value=task_type)
        for col in range(1, 10):
            cell = ws4.cell(row=row_num, column=col)
            cell.border = THIN_BORDER
            cell.font = DATA_FONT
            cell.alignment = CENTER if col != 4 and col != 6 and col != 9 else LEFT
        row_num += 1

# 额外空行
for row in range(row_num, row_num + 20):
    ws4.cell(row=row, column=1, value=row - 3)
    for col in range(1, 10):
        cell = ws4.cell(row=row, column=col)
        cell.border = THIN_BORDER
        cell.font = DATA_FONT
        cell.alignment = CENTER

col_widths_4 = [6, 18, 14, 30, 12, 28, 12, 12, 20]
for i, w in enumerate(col_widths_4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

dv_repeat = DataValidation(type='list', formula1='"可重复,不可重复,每日1次,每周1次,限时1次"', allow_blank=True)
dv_repeat.add(f'G4:G100')
ws4.add_data_validation(dv_repeat)

dv_task_type = DataValidation(type='list', formula1='"日常任务,激励广告任务,社交任务,成长任务,成就任务,活动任务,新手任务"', allow_blank=True)
dv_task_type.add('C4:C100')
ws4.add_data_validation(dv_task_type)

ws4.freeze_panes = 'D4'

# ===== Sheet 5: 经济模型计算 =====
ws5 = wb.create_sheet('经济模型计算')
ws5.sheet_view.showGridLines = False
ws5.merge_cells('A1:H1')
ws5['A1'] = '经济模型 — 关键指标计算'
ws5['A1'].font = TITLE_FONT
ws5['A1'].fill = TITLE_FILL
ws5['A1'].alignment = CENTER
ws5.row_dimensions[1].height = 35

# 说明
ws5.merge_cells('A2:H2')
ws5['A2'] = '填入绿色单元格数据，白色单元格自动计算 ｜ 核心公式：白嫖集数 = 每日产出上限 / 解锁1集所需金币'
ws5['A2'].font = NOTE_FONT
ws5['A2'].alignment = LEFT

econ_headers = [
    '序号', '产品名称',
    '每日产出上限\n(金币)', '解锁1集\n所需金币',
    '理论白嫖\n集数/天', '广告eCPM\n($)',
    '单集广告\n收入($)', '单集ROAS\n(广告/解锁)'
]

for col, h in enumerate(econ_headers, 1):
    cell = ws5.cell(row=3, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

ws5.row_dimensions[3].height = 40

for row in range(4, 24):
    ws5.cell(row=row, column=1, value=row - 3)
    for col in range(1, 9):
        cell = ws5.cell(row=row, column=col)
        cell.border = THIN_BORDER
        cell.font = DATA_FONT
        cell.alignment = CENTER

    # 自动计算列：E=白嫖集数(C/D), H=ROAS(G/D)
    # C=每日产出上限(绿色输入), D=解锁1集金币(绿色输入), F=eCPM(绿色输入), G=单集广告收入(绿色输入)
    for col in [3, 4, 6, 7]:
        ws5.cell(row=row, column=col).fill = HIGHLIGHT_FILL

    # E列 = C/D 白嫖集数
    ws5.cell(row=row, column=5).value = f'=IF(AND(C{row}>0,D{row}>0),ROUND(C{row}/D{row},1),"")'
    ws5.cell(row=row, column=5).number_format = '0.0'
    # H列 = G/D ROAS
    ws5.cell(row=row, column=8).value = f'=IF(AND(G{row}>0,D{row}>0),ROUND(G{row}/D{row},2),"")'
    ws5.cell(row=row, column=8).number_format = '0.00'

col_widths_5 = [6, 20, 16, 14, 14, 14, 14, 14]
for i, w in enumerate(col_widths_5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

ws5.freeze_panes = 'C4'

# ===== Sheet 6: 变现路径设计 =====
ws6 = wb.create_sheet('变现路径设计')
ws6.sheet_view.showGridLines = False
ws6.merge_cells('A1:K1')
ws6['A1'] = '变现路径 — 广告+内购+提现设计对比'
ws6['A1'].font = TITLE_FONT
ws6['A1'].fill = TITLE_FILL
ws6['A1'].alignment = CENTER
ws6.row_dimensions[1].height = 35

monetize_headers = [
    '序号', '产品名称',
    '广告变现\n点位数量', '激励视频\n频率(次/日)',
    '插屏广告\n频率', '原生广告\n位置',
    '内购金币\n档位(最少)', '内购金币\n档位(最多)',
    '1$=\n多少金币', '提现\n门槛($)',
    '提现方式\n(PayPal等)'
]

for col, h in enumerate(monetize_headers, 1):
    cell = ws6.cell(row=2, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

ws6.row_dimensions[2].height = 40

for row in range(3, 53):
    ws6.cell(row=row, column=1, value=row - 2)
    for col in range(1, 12):
        cell = ws6.cell(row=row, column=col)
        cell.border = THIN_BORDER
        cell.font = DATA_FONT
        cell.alignment = CENTER

col_widths_6 = [6, 20, 14, 14, 12, 16, 14, 14, 12, 12, 16]
for i, w in enumerate(col_widths_6, 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

ws6.freeze_panes = 'C3'

# ===== Sheet 7: 典型样本深度拆解 =====
ws7 = wb.create_sheet('典型样本深度拆解')
ws7.sheet_view.showGridLines = False

deep_sections = [
    ('产品基本信息', [
        ('产品名称', ''),
        ('厂商', ''),
        ('上线时间', ''),
        ('主要市场', ''),
        ('变现模式', ''),
        ('DAU预估', ''),
    ]),
    ('金币系统总览', [
        ('金币名称', ''),
        ('金币获取入口数量', ''),
        ('金币消耗场景数量', ''),
        ('每日产出上限', ''),
        ('解锁1集所需金币', ''),
        ('理论白嫖集数/天', ''),
        ('是否支持提现', ''),
        ('提现门槛', ''),
    ]),
    ('获取入口清单', [
        ('入口1: 签到', '奖励: __ / 频次: __ / 递增: __'),
        ('入口2: 激励视频', '奖励: __ / 频次: __ / 上限: __'),
        ('入口3: 分享', '奖励: __ / 频次: __ / 上限: __'),
        ('入口4: 邀请', '奖励: __/人 / 上限: __'),
        ('入口5: 追剧任务', '奖励: __ / 条件: __'),
        ('入口6: 成就', '奖励: __ / 条件: __'),
        ('入口7: 活动/其他', '奖励: __ / 条件: __'),
    ]),
    ('消耗场景清单', [
        ('场景1: 解锁剧集', '单价: __ / 阶梯价: __'),
        ('场景2: 抽奖/转盘', '单价: __ / 奖池: __'),
        ('场景3: 虚拟道具', '价格区间: __'),
        ('场景4: 其他', '说明: __'),
    ]),
    ('任务系统完整清单', [
        ('日常-签到', '奖励: __ / 刷新: 每日'),
        ('日常-观看N集', '奖励: __ / 门槛: __集'),
        ('广告-看N个视频', '奖励: __ / 门槛: __个'),
        ('社交-邀请好友', '奖励: __/人 / 上限: __人'),
        ('社交-分享', '奖励: __ / 平台: __'),
        ('成长-累计观看', '奖励: __ / 里程碑: __'),
        ('成就-首次充值', '奖励: __'),
        ('成就-连续登录', '奖励: __ / 天数: __'),
        ('活动-限时', '奖励: __ / 时间: __'),
        ('新手-引导任务', '奖励: __ / 步骤: __'),
    ]),
    ('广告点位清单', [
        ('激励视频-任务奖励', '位置: __ / 频率: __'),
        ('激励视频-解锁剧集', '位置: __ / 频率: __'),
        ('插屏广告', '位置: __ / 频率: __'),
        ('原生广告', '位置: __'),
        ('开屏广告', '有/无'),
    ]),
    ('经济模型分析', [
        ('每日总产出(金币)', ''),
        ('每日总消耗(金币/活跃用户)', ''),
        ('净产出/消耗比', ''),
        ('广告eCPM($)', ''),
        ('单集广告收入($)', ''),
        ('单集ROAS', ''),
        ('金币通胀/通缩判断', ''),
    ]),
    ('亮点与借鉴', [
        ('设计亮点1', ''),
        ('设计亮点2', ''),
        ('设计亮点3', ''),
        ('可借鉴的点', ''),
        ('需要避免的问题', ''),
    ]),
]

# 标题行
ws7.merge_cells('A1:D1')
ws7['A1'] = '典型样本深度拆解模板（每个产品复制一份Sheet）'
ws7['A1'].font = TITLE_FONT
ws7['A1'].fill = TITLE_FILL
ws7['A1'].alignment = CENTER
ws7.row_dimensions[1].height = 35

row = 3
for section_name, items in deep_sections:
    # Section header
    ws7.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws7.cell(row=row, column=1, value=section_name)
    cell.font = SUB_HEADER_FONT
    cell.fill = SUB_HEADER_FILL
    cell.alignment = LEFT
    cell.border = THIN_BORDER
    row += 1

    for label, default_val in items:
        ws7.cell(row=row, column=1, value=label)
        ws7.cell(row=row, column=1).font = DATA_FONT
        ws7.cell(row=row, column=1).border = THIN_BORDER
        ws7.cell(row=row, column=1).alignment = LEFT

        ws7.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws7.cell(row=row, column=2, value=default_val)
        ws7.cell(row=row, column=2).font = DATA_FONT
        ws7.cell(row=row, column=2).border = THIN_BORDER
        ws7.cell(row=row, column=2).alignment = LEFT

        if default_val:
            ws7.cell(row=row, column=2).fill = HIGHLIGHT_FILL
        row += 1
    row += 1  # 空行分隔

ws7.column_dimensions['A'].width = 28
ws7.column_dimensions['B'].width = 20
ws7.column_dimensions['C'].width = 20
ws7.column_dimensions['D'].width = 20

# ===== 保存 =====
output_path = '海外短剧金币系统分析模板.xlsx'
wb.save(output_path)
print(f'Excel saved to: {output_path}')
